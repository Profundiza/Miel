from django.http import HttpResponse
from openpyxl import Workbook

from .models import *


def export(request):
    # each item is a line in the excel sheet
    # it can be anything you pull from the db
    rest_id = request.user.restaurante.id
    wb = Workbook(write_only=True)

    proveedores = Proveedor.objects.filter(restaurante_id=rest_id)
    ws_prov = wb.create_sheet(title='Proveedores')
    ws_prov.append(['Nombre',
                    'Telefono',
                    'Representante',
                    'Telefono de Representante',
                    'Correo Electronico'])
    for proveedor in proveedores:
        ws_prov.append([
            proveedor.nombre,
            str(proveedor.telefono),
            proveedor.representante,
            str(proveedor.telefono_de_representante),
            proveedor.correo_electronico])

    ingredientes = Ingrediente.objects.filter(restaurante_id=rest_id)
    ws_ing = wb.create_sheet(title='Ingredientes')
    ws_ing.append([
        'Nombre',
        'Marca',
        'Proveedor',
        'Costo ($)',
        'Medida',
        'Cantidad',
        'Costo Unitario ($)'])
    for ing in ingredientes:
        ws_ing.append([
            ing.nombre,
            ing.marca,
            ing.proveedor.nombre,
            ing.costo,
            ing.medida,
            ing.cantidad,
            ing.unit_cost
        ])

    recetas = Receta.objects.filter(restaurante_id=rest_id)
    ws_rec = wb.create_sheet(title='Recetas')
    ws_rec.append([
        'Nombre',
        'Costo ($)',
        'Cantidad',
        'Medida',
        'Costo Unitario ($)'
    ])
    for rec in recetas:
        ws_rec.append([
            rec.nombre,
            rec.costo,
            rec.cantidad,
            rec.medida,
            rec.unit_cost
        ])

    platillos = Platillo.objects.filter(restaurante_id=rest_id, bebida=False)
    ws_plat = wb.create_sheet(title='Platillos')
    ws_plat.append([
        'Nombre',
        'Costo ($)',
        'Precio ($)',
        'Costo / Precio (%)',
        'Ganancia ($)'
    ])
    for plat in platillos:
        ws_plat.append([
            plat.nombre,
            plat.costo,
            plat.precio,
            plat.costo_percentaje,
            plat.ganancia
        ])

    bebidas = Platillo.objects.filter(restaurante_id=rest_id, bebida=True)
    ws_beb = wb.create_sheet(title='Bebidas')
    ws_beb.append([
        'Nombre',
        'Costo ($)',
        'Precio ($)',
        'Costo / Precio (%)',
        'Ganancia ($)'
    ])
    for beb in bebidas:
        ws_plat.append([
            beb.nombre,
            beb.costo,
            beb.precio,
            beb.costo_percentaje,
            beb.ganancia
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=miel-tech.xlsx'

    wb.save(response)

    return response
